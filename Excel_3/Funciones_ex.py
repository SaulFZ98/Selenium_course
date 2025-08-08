import openpyxl
from openpyxl.workbook import Workbook


class Fun_excel():
    def __init__(self,driver):
        self.driver = driver

    def getRowCount(file,path,sheetName):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return (sheet.max_row)

    def getColumCount(file,sheetName):
        Workbook = openpyxl.load_workbook(file)
        sheet = Workbook[sheetName]
        return (sheet.max_column)

    def readData(file,path,sheetName,rownum,columnum):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row=rownum, column=columnum).value

    def writeData(file,path,sheetName,rownum,columnum,data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row=rownum,column=columnum).value = data
        Workbook.save(path)
